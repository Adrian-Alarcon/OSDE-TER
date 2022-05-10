import rutas
# from txva01 import va01
from va01_2 import va01_2
from zsd_toma import toma
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from getpass import getuser
import shutil
import openpyxl
from shutil import copy
import shutil
import pythoncom
from lecturaPdf import lectorPdf
import time
from openpyxl import load_workbook


def interfaz():

     user = getuser()
    
     root = Tk()
     
     root.title("FACTURADOR_CASA")
     root.resizable(0,0)
     root.geometry('300x300+500+50'.format(500, 600))

     miFrame=Frame(root,width=500)
     miFrame.pack()
     miFrame2=Frame(root)
     miFrame2.pack()


     def facturador():

          pythoncom.CoInitialize()
          # Crear una copia del excel padre.
          shutil.copy(rutas.archivo_excel, rutas.archivo_excel_trabajo)

          #---------VARIABLES--------#
          afiliado_anterior = None
          afiliado_actual = None

          l_descripciones = []
          l_id_productos = []
          l_cantidades_va01 = []
          l_afiliados_sap = []
          l_dispones = []
          l_canales = []
          l_sectores = []
          l_ped_ext = []
          l_convenios = []

          l_descripciones_facturar = []
          l_mat_facturar = []
          l_cant_facturar = []
          fecha_entrega = []
          filas_completar = []
          fila = 9
          l_filas = []
          #---------------------------#
          
          excel_trabajo = load_workbook(rutas.archivo_excel_trabajo, data_only=True)
          excel_trabajo = openpyxl.load_workbook(filename=rutas.archivo_excel_trabajo, data_only=True)
          try:
               h_t = excel_trabajo["inicio"]
               cont = 9
               max_filas = cont
               
               # CALCULAR LA CANTIDAD DE FILAS MAXIMAS #
               while h_t[f"A{max_filas}"].value != None:
                    max_filas += 1
          

               for i in range(9, max_filas):
                    revision = h_t[f"F{i}"].value
                    
                    if revision == "SI":
                         print(f"AFILIADO {afiliado_actual} para REVISION")
                         continue
                    else:
                         l_afiliados_sap.append(h_t[f"M{i}"].value)
                         l_id_productos.append(h_t[f"N{i}"].value)
                         l_descripciones.append(h_t[f"A{i}"].value)
                         l_cantidades_va01.append(h_t[f"D{i}"].value)
                         l_canales.append(h_t[f"W{i}"].value)
                         l_sectores.append(h_t[f"X{i}"].value)
                         l_ped_ext.append(h_t[f"E{i}"].value)
                         l_dispones.append(h_t[f"L{i}"].value)
                         fecha_entrega.append(h_t[f"T{i}"].value)
                         l_convenios.append(h_t[f"O{i}"].value)
                         filas_completar.append(str(i))
               print(f"Filas a completar:", filas_completar)

               # -> FACTURACION <--
               for i in range(len(l_afiliados_sap)):
                    afiliado_actual = l_afiliados_sap[i]

                    
                    if afiliado_actual == afiliado_anterior or afiliado_anterior == None:
                         print(f"Agregando Materiales: {l_descripciones[i]} | AFILIADO:{afiliado_actual} |")
                         l_descripciones_facturar.append(l_descripciones[i])
                         l_mat_facturar.append(l_id_productos[i])
                         l_cant_facturar.append(l_cantidades_va01[i])
                         l_filas.append(filas_completar[i])
                    
                    elif afiliado_actual != afiliado_anterior:
                         print(f"Afiliado Diferente")
                         print(f" ------ SE FACTURA AF: {afiliado_anterior} ------ ")
                         print(f"\tSE FACTURA MATERIALES: {l_mat_facturar}\n\n")
                         pedidova01 = va01_2(0, l_canales[i-1], l_sectores[i-1], l_ped_ext[i-1], l_dispones[i-1], fecha_entrega[i-1], l_mat_facturar, l_cant_facturar, l_convenios[i-1], l_descripciones_facturar)
                         time.sleep(1)
                         _toma = toma(0, pedidova01, l_dispones[i-1], afiliado_anterior, l_canales[i-1])

                         # Completar Excel con pedido generado
                         for fila in l_filas:
                              h_t[f"Y{fila}"] = _toma

                         l_descripciones_facturar.clear()
                         l_mat_facturar.clear()
                         l_cant_facturar.clear()
                         l_filas.clear()

                         print(f"Agregando Materiales: {l_descripciones[i]} | AFILIADO:{afiliado_actual} |")
                         l_descripciones_facturar.append(l_descripciones[i])
                         l_mat_facturar.append(l_id_productos[i])
                         l_cant_facturar.append(l_cantidades_va01[i])
                         l_filas.append(filas_completar[i])

                    if i == len(l_afiliados_sap) - 1:
                         print(f" ------ ULTIMO AFILIADO: {afiliado_actual}")
                         print(f"\tSE FACTURA MATERIALES: {l_mat_facturar}")
                         pedidova01 = va01_2(0, l_canales[i], l_sectores[i], l_ped_ext[i], l_dispones[i], fecha_entrega[i], l_mat_facturar, l_cant_facturar, l_convenios[i], l_descripciones_facturar)
                         time.sleep(1)
                         _toma = toma(0, pedidova01, l_dispones[i], afiliado_actual, l_canales[i])

                         # Completar Excel con pedido generado
                         for fila in l_filas:
                              h_t[f"Y{fila}"] = _toma

                         break
                    
                    afiliado_anterior = afiliado_actual


          except Exception as e:
               print(f"Excepcion el Excel de Trabajo {e}")
          finally:
               excel_trabajo.save(rutas.archivo_excel_trabajo)
               excel_trabajo.close()
     
     
     def lecturaPdfs():
          resultado_lectura = lectorPdf()
          print(resultado_lectura)

     
     botonPdf = Button(miFrame, text="Leer PDFS", command=lecturaPdfs)
     botonCrear = Button(miFrame, text="Ejecutar", command=facturador)
     botonPdf.grid(row = 10, column = 2, sticky = "e", padx = 10, pady = 10)
     botonCrear.grid(row = 13, column = 2, sticky = "e", padx = 15, pady = 10)

     root.mainloop()

interfaz()